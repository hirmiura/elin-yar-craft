#!/usr/bin/env -S python3
# SPDX-License-Identifier: MIT
# Copyright 2024 hirmiura (https://github.com/hirmiura)

from __future__ import annotations

import argparse
import copy
import csv
import re
import sys
import warnings
from typing import Any, ClassVar

import openpyxl
from pydantic import BaseModel


def procee_args() -> argparse.Namespace:
    """コマンドライン引数を処理する

    Returns:
        argparse.Namespace: 処理した引数
    """
    parser = argparse.ArgumentParser(description="入力ファイルを翻訳する")
    parser.add_argument(
        "-i",
        metavar="IN",
        nargs="?",
        type=argparse.FileType("r", encoding="utf-8-sig"),  # BOM付いてるやん!!
        default=sys.stdin,
        help="入力ファイル。未指定時は標準入力",
    )
    parser.add_argument(
        "-o",
        metavar="OUT",
        nargs="?",
        type=argparse.FileType("w"),
        default=sys.stdout,
        help="出力ファイル。未指定時は標準出力",
    )
    parser.add_argument(
        "-t", metavar="TRANS", dest="trans", required=True, help="翻訳に使うThing.xlsxファイル"
    )
    lang_choices = ["cn"]
    parser.add_argument(
        "-l", metavar="LANG", dest="lang", required=True, choices=lang_choices, help="対象の言語"
    )
    parser.add_argument("--version", action="version", version="%(prog)s 0.1.0")
    args = parser.parse_args()

    return args


class TransThing(BaseModel):
    key_id: ClassVar[str] = "id"
    key_name: ClassVar[str] = "name"
    key_name_EN: ClassVar[str] = "name_EN"  # noqa: N815
    key_detail: ClassVar[str] = "detail"
    key_unit: ClassVar[str] = "unit"
    key_unknown: ClassVar[str] = "unknown"
    key_roomName: ClassVar[str] = "roomName"  # noqa: N815
    key_name2: ClassVar[str] = "name2"
    trans_list: list[dict[str, Any]] = []

    @classmethod
    def load(cls, file: str) -> TransThing:
        # オブジェクト生成
        self = cls()
        # ファイル読み込み
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(file)
        ws = wb["Thing"]
        # 1行目(ヘッダ)
        header = ws[1]
        # 3行目以降(2行目はゴミ)
        for row in ws.iter_rows(min_row=3):
            row_dic = {}
            for k, v in zip(header, row):
                row_dic[k.value] = v.value
            self.trans_list.append(row_dic)
        return self

    def translate(self, lang: str, line: dict[str, Any]) -> dict[str, Any] | None:
        for trans in self.trans_list:
            trans_id: str = trans[TransThing.key_id]
            line_id: str = line[TransThing.key_id]
            pure_line_id = re.sub(r"^YarCraft_", "", line_id)
            pure_line_id = re.sub(r"_(wood|stone|metal|cloth)?(_q\d)?$", "", pure_line_id)
            if (  # idが一致するか
                trans_id.casefold() == pure_line_id.casefold()
            ):
                new = copy.deepcopy(line)
                for key in [
                    TransThing.key_name,
                    TransThing.key_detail,
                    TransThing.key_unit,
                    TransThing.key_unknown,
                    TransThing.key_roomName,
                ]:
                    new[key] = trans[key]
                    if key == TransThing.key_name:
                        new[key] = get_key_name(lang, new[TransThing.key_id], new[key])

                return new
        return None


def get_key_name(lang: str, id: str, name: str) -> str:
    # 末尾が _q\d か?
    suffix_match = re.search(r"(_q\d)$", id)
    if not suffix_match:
        return name  # マッチしなければリターン

    # ランク分け
    new_name: str = name + " "
    match suffix_match[0]:
        case "_q1":
            match lang:
                case "cn":
                    new_name += "优质品"
                case _:
                    new_name += "good"
        case "_q2":
            match lang:
                case "cn":
                    new_name += "奇迹"
                case _:
                    new_name += "miracle"
        case "_q3":
            match lang:
                case "cn":
                    new_name += "神器"
                case _:
                    new_name += "godly"
        case "_q4":
            match lang:
                case "cn":
                    new_name += "特制"
                case _:
                    new_name += "special"
    return new_name


def main() -> int:
    """メイン関数

    Returns:
        int: 成功時は0を返す
    """
    args = procee_args()

    # 翻訳ファイルの読み込み
    trans = TransThing.load(args.trans)

    # 入力csvファイルの読み込み
    reader = csv.DictReader(args.i)
    data = list(reader)
    header = data[0].keys()

    # 生成
    result = []
    for line in data:
        translated = trans.translate(args.lang, line)
        if translated:
            result.append(translated)
        else:
            print(f"Warning: Not matched. {line['id']}", file=sys.stderr)

    # 書き出し
    writer = csv.DictWriter(args.o, header)
    writer.writeheader()
    writer.writerows(result)

    return 0


if __name__ == "__main__":
    sys.exit(main())
